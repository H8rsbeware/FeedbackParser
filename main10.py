# openpyxl reads, writes, and saves .xlsx and .xls
import openpyxl as oxl
# same base library as openpyxl, but built for .ods
from pyexcel_ods import get_data as calc
# Json allows me to print formatted dictonaries to collect raw data
import json
# OS allows for system management, like grabbing files
from os import listdir
from os.path import isfile, join


# Constants
PATH = "/Example-Input"
RESULT = "/Example-Output"
# Holds the row ranges for searching spreadsheets
CLASSRANGE = [1,12]
LEADERRANGE = [16,19]

# Stores bulk data from file reads
level5 = []
level6 = []

# The classes act like a standardised data structure, making it easier to handle data
class person:
    def __init__(self, name, rel, team, crt, prod, pos, neg):
        self.name = name
        self.rel = rel
        self.team = team
        self.crt = crt
        self.prod = prod
        self.pos = pos
        self.neg = neg

    # formatting print command
    def __str__(self):
        return f"{self.name},{self.rel},{self.team},{self.crt},{self.prod},positive: {self.pos}, negative: {self.neg}"

    # Returns all values in an array, with no formatting
    def getCSV(self):
        return [self.name, self.rel,self.team,self.crt,self.prod,self.pos,self.neg]

# Leader class is kind of redundant, could of been condensed into one class
class leader:
    def __init__(self, name, mot, fair, cpt, cmt, pos, neg):
        self.name = name
        self.mot = mot
        self.fair = fair
        self.cpt = cpt
        self.cmt = cmt
        self.pos = pos
        self.neg = neg

    def __str__(self):
        return f"{self.name},{self.mot},{self.fair},{self.cpt},{self.cmt},positive: {self.pos}, negative: {self.neg}"

    def getCSV(self):
        return [self.name, self.mot, self.fair, self.cpt, self.cmt, self.pos, self.neg]



def main():
    # grabs file names from a path
    files = get_files(PATH)
    amountOfFiles = len(files)

    # for each feedback form
    for file in files:
        # store the file type i.e. xls, xlsx, csv, or ods
        extension = file.split(".")[1]

        # if its a standard excel format
        if extension == "xls" or extension == "xlsx":
            # Converts rows from spreadsheets into the classes above, adds it to level5/level6 array
            convert_xsl_dict(file)
        # if its an open standard
        elif extension == "ods":
            # Same as XLS, just with a different library
            convert_ods_dict(file)
        else:
            throwErr("format", file)

    # Dictonaries to hold sorted values
    level5sort = {
    }
    level6sort = {

    }

    # condenses the raw data into 1 instance per person
    for entry in level5:
        level5sort = condense_dict(level5sort, entry)
    for entry in level6:
        level6sort = condense_dict(level6sort, entry)

    # writes the raw results to json (thanks python)
    f = open(f"{RESULT}/rawResults.json","a")
    f.write(json.dumps(level5sort, indent=4))
    f.write("\n")
    f.write(json.dumps(level6sort, indent=4))
    f.close()

    # This is about the point in the night where making it neat didnt matter, and functions were a waste of energy
    # Creates a new spreadsheet
    file = oxl.Workbook()
    act = file.active

    act.title = "Averages"

    # Adds headings for the members
    act["A1"] = "Members (AVGS)"
    act["B1"] = "Reliability"
    act["C1"] = "Teamwork"
    act["D1"] = "Creativity"
    act["E1"] = "Productivity"
    act["F1"] = "Total avg/40"

    row = 2
    # writes all of the level 5 results and feedback out
    for member in level5sort:
        file_writer(act, member, level5sort, row)
        row += 1

    row +=1
    # Adds leader headings
    act[f"A{row}"] = "Leaders (AVGS)"
    act[f"B{row}"] = "Motivating"
    act[f"C{row}"] = "Fair"
    act[f"D{row}"] = "Competency"
    act[f"E{row}"] = "Commitment"
    act[f"F{row}"] = "Total avg/40"
    row +=1
    # does the same results and feedback process
    for leader in level6sort:
        file_writer(act, leader, level6sort, row)
        row += 1

    # Saves the files
    file.save(f"{RESULT}/Week-x-Averages.xlsx")


    # Quick check to make sure all data is obtained, should equal 12 and 4 respectively (class amounts)
    print(f"classmates accounted : {len(level5sort)}\nleaders accounted    : {len(level6sort)}\ntotal files     : {len(level5) + len(level6) / amountOfFiles}")

# Gets all scores for a person, averages them, writes then, and then creates a seperate feedback file
def file_writer(writer, entity, dict, row):
    # EXCEL THINGS
    key = str(entity)
    ref = dict[key]
    submissions = len(ref)
    # Averages are calculated (sum of scores in category, divided by responses to that category
    avgC1 = sum_array(return_instances("1", ref)) / submissions
    avgC2 = sum_array(return_instances("2", ref)) / submissions
    avgC3 = sum_array(return_instances("3", ref)) / submissions
    avgC4 = sum_array(return_instances("4", ref)) / submissions
    avgTotal = (avgC1 + avgC2 + avgC3 + avgC4)
    # Writes them to the spreadsheet
    writer[f"A{row}"] = key
    writer[f"B{row}"] = avgC1
    writer[f"C{row}"] = avgC2
    writer[f"D{row}"] = avgC3
    writer[f"E{row}"] = avgC4
    writer[f"F{row}"] = avgTotal

    # POSITIVE AND NEGATIVE POINT HANDLING
    # creates a new file under the persons name
    f = open(f"{RESULT}/{key}-feedback.txt","a")
    pos = return_instances("pos", ref)
    neg = return_instances("neg",ref)
    # Adds all positive feedback
    f.write("What to do more of:\n")
    for point in pos:
        f.write(f"\t{point}\n")
    # Adds all negative feedback
    f.write("What to do less of:\n")
    for point in neg:
        f.write(f"\t{point}\n")
    # Closes the file
    f.close()


# Returns all instances under a dict or sub dict of a key as an array
def return_instances(key, dict):
    res = [sub[str(key)] for sub in dict]
    return res

# Sums all numbers if the array only contains numbers
def sum_array(array):
    s = 0
    try:
        for x in array:
            s += x
        return s

    except TypeError:
        return throwErr(err="type")


def condense_dict(dict, entry):
        # These are poorly named variables, dont look
        e = entry.getCSV()
        n = e[0]
        r = e[1]
        c = e[2]
        t = e[3]
        p = e[4]
        ps = e[5]
        ng = e[6]

        # If the output doesnt already contains an instance of the person
        if dict.get(n) is None:
            # Create a new instance
            dict.update({str(n): [{"1": r, "2": c, "3": t, "4": p, "pos": ps, "neg": ng}]})
        else:
            # if they do already exist, just add the new stats on.
            dict[n].append({"1": r, "2": c, "3": t, "4": p, "pos": ps, "neg": ng})
        return dict


# Both convert_file_dict just take a file type, and change each row into a standard class
def convert_ods_dict(file, inlMaxValue=40, inlMinValue=0):
    ods = calc(f"{PATH}/{file}", data_only=True)
    # for each team member
    for i in range(CLASSRANGE[0], CLASSRANGE[1] + 1):
        # grab their row
        p = ods["Sheet1"][i]
        print(p)
        # make sure their score is correctly inputted or throw an error
        if inlMinValue <= p[5] <= inlMaxValue:
            # create object and add it to stores
            e = person(p[0], p[1], p[2], p[3], p[4], p[6], p[7])
            level5.append(e)
        else:
            level5.append({"name":p[0], "error":"wrong values"})
            throwErr("total", f"{file}:{p[0]}")

    # works the same as above just using a different object for ease later

    for i in range(LEADERRANGE[0]-1, LEADERRANGE[1]):
        l = ods["Sheet1"][i]
        print(l)
        if inlMinValue <= l[5] <= inlMaxValue:
            e = leader(l[0], l[1], l[2], l[3], l[4],l[6],l[7])
            level6.append(e)
        else:

            level6.append({"name": l[0], "error": "wrong values"})
            throwErr("total", f"{file}:{l[0]}")



def convert_xsl_dict(file, inlMaxValue=40, inlMinValue=0):
    xsl = oxl.load_workbook(PATH + "/" + file)
    sht = xsl.active
    for row in sht.iter_rows(min_row=CLASSRANGE[0]+1, max_row=CLASSRANGE[1]+1, min_col=0, max_col=10):
        n = row[0].value
        r = row[1].value
        t = row[2].value
        c = row[3].value
        p = row[4].value
        ps = row[6].value
        ng = row[7].value

        
        if inlMinValue <= r + t + c + p <= inlMaxValue:
            e = person(n, r, t, c, p, ps, ng)
            level5.append(e)
        else:
            if n != "Members":
                level5.append({"name": n, "error": "wrong values"})
            throwErr("total", f"{file}:{n}")

    for row in sht.iter_rows(min_row=LEADERRANGE[0]+1, max_row=LEADERRANGE[1]+1, min_col=0, max_col=10):
        n = row[0].value
        r = row[1].value
        t = row[2].value
        c = row[3].value
        p = row[4].value
        ps = row[6].value
        ng = row[7].value

        if inlMinValue <= r + t + c + p <= inlMaxValue:
            e = leader(n, r, t, c, p, ps, ng)
            level6.append(e)
        else:

            level6.append({"name": n, "error": "wrong values"})
            throwErr("total", f"{file}:{n}")



# returns a list of file names from a path
def get_files(directory):
    files = [f for f in listdir(directory) if isfile(join(directory, f))]
    return files

# Not fatal failer handler.
def throwErr(err = "", loc = ""):
    if err == "total":
        return print(f"ERROR: {loc} has not been totalled correctly")
    if err == "format":
        return print(f"ERROR: {loc} is not formatted correctly, only allows xslx, xsl, and ods")
    if err == "type":
        return print("ERROR: theres text in one of the cells or something. wtf?")
    return print("UNHANDLED: im pretty sure this wont ever get triggered")



if __name__ == '__main__':
    main()


# Inputs : 
#   xlsx, xls, ods, csv
#   Conversion to standard row class (name, [row values], [[row strings])
#   selectable Row ranges, catagories, column ranges, checks, spreadsheet rules.
# Manipulations : 
#   Averages - Mean Mode Median
#   Graphs - Bar Pie Scatter 
#       Identify best maybe
#   Organise, Split, and manipulate string values, and create txt files containing them
# Outputs :
#   Excel sheets, ODS sheets, Txt files, Json

# !FUTURE PLANS!
#   Form/File creator for input and output forms
#   JS local front end for inputing data forms and creating them 
#               (ps, i suggest looking at the json output on the repo, it will be useful for raw data.)
#   Graph visualisation and selection on app
#   Independence but support for excel and ods
