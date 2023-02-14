# openpyxl reads, writes, and saves .xlsx and .xls
import openpyxl as oxl
# same base library as openpyxl, but built for .ods
from pyexcel_ods import get_data as calc
# Json allows me to print formatted dictonaries to collect raw data
import json
# OS allows for system management, like grabbing files
from os import listdir
from os.path import isfile, join

import spreadsheet as st

# Constants
PATH = "/home/user/Documents/UnweightedFeedback"
RESULT = "/home/user/Documents/FeedbackOutput"
RANGES = [[2,13],[16,19]]

VALUES = 4
STRINGS = 2

# Stores bulk data from file reads
level5 = []
level6 = []

# Class as a datastructure for rows
class cRow:
    def __init__(self, key, values=[], strings=[]):
        self.key = key
        self.values = values
        self.strings = strings

    def __str__(self):
        return f"{self.key} : \n\tvalues:{str(self.values)}\n\tstrings:{str(self.strings)}"

    def getCSV(self):
        return [self.key, self.values, self.strings]

def main():
    # grabs file names from a path
    files = get_files(PATH)
    amountOfFiles = len(files)

    # for each feedback form
    for file in files:
        # store the file type i.e. xls, xlsx, csv, or ods
        extension = file.split(".")[1]

        # if its a standard excel format
        if extension == "xls" or extension == "xlsx":
            # Converts rows from spreadsheets into the classes above, adds it to level5/level6 array
            convert_xsl_dict(file)
        # if its an open standard
        elif extension == "ods":
            # Same as XLS, just with a different library
            convert_ods_dict(file)
        else:
            throwErr("format", file)

    # Dictonaries to hold sorted values
    level5sort = {}
    level6sort = {}
    # condenses the raw data into 1 instance per person
    for entry in level5:
        level5sort = condense_dict(level5sort, entry)
    for entry in level6:
        level6sort = condense_dict(level6sort, entry)

    # writes the raw results to json (thanks python)
    f = open(f"{RESULT}/rawResults.json","a")
    f.write(json.dumps(level5sort, indent=4))
    f.write("\n")
    f.write(json.dumps(level6sort, indent=4))
    f.close()

    # This is about the point in the night where making it neat didnt matter, and functions were a waste of energy
    # Creates a new spreadsheet
    file = oxl.Workbook()
    act = file.active

    act.title = "Averages"

    # Adds headings for the members
    act["A1"] = "Members (AVGS)"
    act["B1"] = "Reliability"
    act["C1"] = "Teamwork"
    act["D1"] = "Creativity"
    act["E1"] = "Productivity"
    act["F1"] = "Total avg/40"

    row = 2
    # writes all the level 5 results and feedback out
    for member in level5sort:
        file_writer(act, member, level5sort, row)
        row += 1

    row +=1
    # Adds leader headings
    act[f"A{row}"] = "Leaders (AVGS)"
    act[f"B{row}"] = "Motivating"
    act[f"C{row}"] = "Fair"
    act[f"D{row}"] = "Competency"
    act[f"E{row}"] = "Commitment"
    act[f"F{row}"] = "Total avg/40"
    row +=1
    # does the same results and feedback process
    for leader in level6sort:
        file_writer(act, leader, level6sort, row)
        row += 1

    # Saves the files
    file.save(f"{RESULT}/Week-x-Averages.xlsx")

    # Quick check to make sure all data is obtained, should equal 12 and 4 respectively (class amounts)
    print(f"classmates accounted : {len(level5sort)}\nleaders accounted    : {len(level6sort)}\ntotal files     : {len(level5) + len(level6) / amountOfFiles}")

# Gets all scores for a person, averages them, writes then, and then creates a seperate feedback file
def file_writer(writer, entity, dict, row, ):

    mean = average_mean_row_format(str(entity), dict)
    print(mean)
    i = 1
    # for all average/return values for a key
    for value in mean:
        # converts the int into a char, adds the value to that location on the spreadsheet
        writer[f"{st.simple_i2a(i)}{row}"] = value
        i += 1

    anon_form_format(str(entity), dict, ["What to keep doing:\n", "What to do less:\n"])


# finds the average of column values for a key
def average_mean_row_format(key:str, sorted:dict) -> []:
    i = 1
    ref = sorted[key]
    total = 0
    format = []
    format.append(key)
    while i < VALUES+1:
        avgVal = sum_array(return_instances(f"v{i}", ref)) / len(ref)
        total += avgVal
        format.append(avgVal)
        i += 1
    format.append(total)
    return format


# splits string feedback into categories for a key
def anon_form_format(key:str, sorted:dict, headers:[str]):
    strings = []
    ref = sorted[key]
    i = 1

    f = open(f"{RESULT}/{key}--feedback.txt","a")

    while i < STRINGS+1:
        category = return_instances(f"s{i}", ref)
        f.write(headers[i-1] or f"category {i}")
        for point in category:
            f.write(f"\t{point}\n")
        i += 1
    f.close()


# Returns all instances under a dict or sub dict of a key as an array
def return_instances(key, dict):
    res = [sub[str(key)] for sub in dict]
    return res

# Sums all numbers if the array only contains numbers
def sum_array(array):
    s = 0
    try:
        for x in array:
            s += x
        return s

    except TypeError:
        return throwErr(err="type")

def condense_dict(dict, entry):
        # These are poorly named variables, dont look
        e = entry.getCSV()
        key = e[0]
        values = e[1]
        strings = e[2]

        tempDict = {}
        i = 1
        for value in values:
            tempDict[f"v{i}"] = value
            i += 1
        i = 1
        for rope in strings:
            tempDict[f"s{i}"] = rope
            i += 1
            
        # If the output doesnt already contains an instance of the person
        if dict.get(key) is None:
            # Create a new instance
            dict.update({str(key): [tempDict]})
        else:
            # if they do already exist, just add the new stats on.
            dict[key].append(tempDict)
        return dict

# Both convert_file_dict just take a file type, and change each row into a standard class
def convert_ods_dict(file, inlMaxValue=40, inlMinValue=0):
    ods = calc(f"{PATH}/{file}", data_only=True)
    # For each range in ranges
    for ran in RANGES:
        # Go through each row in that range
        for interval in range(ran[0]-1,ran[1]):
            # read the row, check if the values are right, and add it to the right dictionary
            currentRow = ods["Sheet1"][interval]
            e = ""
            if inlMinValue <= currentRow[5] <= inlMaxValue:
                e = cRow(currentRow[0], [currentRow[1], currentRow[2], currentRow[3], currentRow[4]], [currentRow[6], currentRow[7]])
            else:
                e = {"name": currentRow[0], "error": "wrong values"}
                throwErr("total", f"{file}:{currentRow[0]}")
            # automate headings later
            if ran == RANGES[1]:
                level6.append(e)
            else:
                level5.append(e)            

def convert_xsl_dict(file, inlMaxValue=40, inlMinValue=0):
    xsl = oxl.load_workbook(PATH + "/" + file)
    sht = xsl.active

    for ran in RANGES:
        check = 0
        if ran == RANGES[1]:
            skew = [1,1]
            check = 1
        else:
            skew = [0,0]

        for row in sht.iter_rows(min_row=ran[0]+skew[0], max_row=ran[1]+skew[1], min_col=0, max_col=10):
            key = row[0].value
            values = [row[1].value, row[2].value, row[3].value, row[4].value]
            strings = [row[6].value, row[7].value]
            if inlMinValue <= sum_array(values) <= inlMaxValue:
                e = cRow(key, values, strings)
            else:
                e = {"name": key, "error": "wrong values"}
                throwErr("total", f"{file}:{key}")
            
            if check == 0:
                level5.append(e)
            else:
                level6.append(e)

# returns a list of file names from a path
def get_files(directory):
    files = [f for f in listdir(directory) if isfile(join(directory, f))]
    return files

# Not fatal failer handler.
def throwErr(err = "", loc = ""):
    if err == "total":
        return print(f"ERROR: {loc} has not been totalled correctly")
    if err == "format":
        return print(f"ERROR: {loc} is not formatted correctly, only allows xslx, xsl, and ods")
    if err == "type":
        return print("ERROR: theres text in one of the cells or something. wtf?")
    return print("UNHANDLED: im pretty sure this wont ever get triggered")


if __name__ == '__main__':
    main()

# Inputs : 
#   xlsx, xls, ods, csv
#   Conversion to standard row class (name, [row values], [[row strings])
#   selectable Row ranges, catagories, column ranges, checks, spreadsheet rules.
# Manipulations : 
#   Averages - Mean Mode Median
#   Graphs - Bar Pie Scatter 
#       Identify best maybe
#   Organise, Split, and manipulate string values, and create txt files containing them
# Outputs :
#   Excel sheets, ODS sheets, Txt files, Json

# !FUTURE PLANS!
#   Form/File creator for input and output forms
#   JS local front end for inputing data forms and creating them 
#               (ps, i suggest looking at the json output on the repo, it will be useful for raw data.)
#   Graph visualisation and selection on app
#   Independence but support for excel and ods
