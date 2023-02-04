# openpyxl for excel worksheet opening
import openpyxl as oxl
import json
# pyexcel_ods for excel and calc open standard opening
from pyexcel_ods import get_data as calc
# csv for creating comma separated value files
import csv
# os.listdir gets files in a directory
from os import listdir
from os.path import isfile, join
from os import remove

# Constants
PATH = "/home/user/Documents/TestFeedback"
RESULT = "/home/user/Documents/FeedbackOutput"
CLASSRANGE = [1,12]
LEADERRANGE = [16,19]
level5 = []
level6 = []



# Classes for storing values (can scroll through)
class person:
    # glorified dictonary
    def __init__(self, name, rel, team, crt, prod, pos, neg):
        self.name = name
        self.rel = rel
        self.team = team
        self.crt = crt
        self.prod = prod
        self.pos = pos
        self.neg = neg

    # formatting print command
    def __str__(self):
        return f"{self.name},{self.rel},{self.team},{self.crt},{self.prod},positive: {self.pos}, negative: {self.neg}"

    def isName(self, name):
        if name == self.name:
            return True
        return False

    def getName(self):
        return self.name

    def getCSV(self):
        return [self.name, self.rel,self.team,self.crt,self.prod,self.pos,self.neg]

class leader:
    def __init__(self, name, mot, fair, cpt, cmt, pos, neg):
        self.name = name
        self.mot = mot
        self.fair = fair
        self.cpt = cpt
        self.cmt = cmt
        self.pos = pos
        self.neg = neg

    def __str__(self):
        return f"{self.name},{self.mot},{self.fair},{self.cpt},{self.cmt},positive: {self.pos}, negative: {self.neg}"

    def isName(self, name):
        if name == self.name:
            return True
        return False

    def getName(self):
        return self.name

    def getCSV(self):
        return [self.name, self.mot, self.fair, self.cpt, self.cmt, self.pos, self.neg]



def main():
    # grabs file names from a path
    files = get_files(PATH)
    amountOfFiles = len(files)
    # sets up value stores

    # for each feedback form
    for file in files:
        # store the file type i.e. xls, xlsx, csv, or ods
        extension = file.split(".")[1]

        # if its a standard excel format
        if extension == "xls" or extension == "xlsx":
            convert_xsl_dict(file)
        # if its an open standard
        elif extension == "ods":
            convert_ods_dict(file)
        else:
            throwErr("format", file)

        # sort level 5s and 6, average them, return feedback.
        # add feedback storage to classes. collect seperately for each person.

    level5sort = {
    }
    level6sort = {

    }

    for entry in level5:
        level5sort = condense_dict(level5sort, entry)
    for entry in level6:
        level6sort = condense_dict(level6sort, entry)
    f = open(f"{RESULT}/rawResults.json","a")
    f.write(json.dumps(level5sort, indent=4))
    f.write("\n")
    f.write(json.dumps(level6sort, indent=4))
    f.close()

    # This is about the point in the night where making it neat didnt matter, and functions were a waste of energy
    file = oxl.Workbook()
    act = file.active

    act.title = "Averages"

    spreadsheetAmount = amountOfFiles

    act["A1"] = "Members (AVGS)"
    act["B1"] = "Reliability"
    act["C1"] = "Teamwork"
    act["D1"] = "Creativity"
    act["E1"] = "Productivity"

    row = 2

    for member in level5sort:
        file_writer(act, member, level5sort, row)
        row += 1

    row +=1
    act[f"A{row}"] = "Leaders (AVGS)"
    act[f"B{row}"] = "Motivating"
    act[f"C{row}"] = "Fair"
    act[f"D{row}"] = "Competency"
    act[f"E{row}"] = "Commitment"
    row +=1

    for leader in level6sort:
        file_writer(act, leader, level6sort, row)
        row += 1


    file.save(f"{RESULT}/Week-x-Averages.xlsx")


    # should equal 12 and 4 respectively (class amounts)
    print(f"classmates accounted : {len(level5sort)}\nleaders accounted    : {len(level6sort)}")

def file_writer(writer, entity, dict, row):
    # EXCEL THINGS
    key = str(entity)
    ref = dict[key]
    submissions = len(ref)
    a1 = sum_array(return_instances("1", ref)) / submissions
    a2 = sum_array(return_instances("2", ref)) / submissions
    a3 = sum_array(return_instances("3", ref)) / submissions
    a4 = sum_array(return_instances("4", ref)) / submissions
    writer[f"A{row}"] = key
    writer[f"B{row}"] = a1
    writer[f"C{row}"] = a2
    writer[f"D{row}"] = a3
    writer[f"E{row}"] = a4
    # POSITIVE AND NEGATIVE POINT HANDLING
    f = open(f"{RESULT}/{key}-feedback.txt","a")
    pos = return_instances("pos", ref)
    neg = return_instances("neg",ref)
    f.write("What to do more of:\n")
    for point in pos:
        f.write(f"\t{point}\n")
    f.write("What to do less of:\n")
    for point in neg:
        f.write(f"\t{point}\n")
    f.close()



def return_instances(key, list):
    res = [sub[str(key)] for sub in list]
    return res

def sum_array(array):
    s = 0
    try:
        for x in array:
            s += x
        return s

    except TypeError:
        return throwErr(err="type")



def condense_dict(dict, entry):

        e = entry.getCSV()
        n = e[0]
        r = e[1]
        c = e[2]
        t = e[3]
        p = e[4]
        ps = e[5]
        ng = e[6]

        if dict.get(n) is None:
            dict.update({str(n): [{"1": r, "2": c, "3": t, "4": p, "pos": ps, "neg": ng}]})
        else:
            dict[n].append({"1": r, "2": c, "3": t, "4": p, "pos": ps, "neg": ng})
        return dict

def convert_ods_dict(file):
    ods = calc(f"{PATH}/{file}")
    # for each team member
    for i in range(CLASSRANGE[0], CLASSRANGE[1] + 1):
        # grab their row
        meCheck = False
        p = ods["Sheet1"][i]
        me = p[9]
        # make sure their score is correctly inputted or throw an error
        if p[6] == 0:
            # create object and add it to stores
            e = person(p[0], p[1], p[2], p[3], p[4], p[7], p[8])
            level5.append(e)
        elif me == 1 and meCheck is False:
            meCheck = True
        else:
            level5.append({"name":p[0], "error":"wrong values"})
            throwErr("total", f"{file}:{p[0]}")

    # works the same as above just using a different object for ease later
    for i in range(LEADERRANGE[0], LEADERRANGE[1] + 1):
        meCheck = False
        l = ods["Sheet1"][i]
        me = l[9]
        if l[6] == 0:
            e = leader(l[0], l[1], l[2], l[3], l[4],l[7],l[8])
            level6.append(e)
        elif me == 1 and meCheck is False:
            meCheck = True
        else:

            level6.append({"name": l[0], "error": "wrong values"})
            throwErr("total", f"{file}:{l[0]}")

def convert_xsl_dict(file):
    xsl = oxl.load_workbook(PATH + "/" + file)
    sht = xsl.active
    for row in sht.iter_rows(min_row=CLASSRANGE[0]+1, max_row=CLASSRANGE[1]+1, min_col=0, max_col=10):
        meCheck = False
        me = row[9].value
        n = row[0].value
        r = row[1].value
        t = row[2].value
        c = row[3].value
        p = row[4].value
        ps = row[7].value
        ng = row[8].value

        if r + t + c + p == 100:
            e = person(n, r, t, c, p, ps, ng)
            level5.append(e)
        elif me == 1 and meCheck is False:
            meCheck = True
        else:
            if n != "Members":
                level5.append({"name": n, "error": "wrong values"})
            throwErr("total", f"{file}:{n}")

    for row in sht.iter_rows(min_row=LEADERRANGE[0]+1, max_row=LEADERRANGE[1]+1, min_col=0, max_col=10):
        meCheck = False
        me = row[9].value
        n = row[0].value
        r = row[1].value
        t = row[2].value
        c = row[3].value
        p = row[4].value
        ps = row[7].value
        ng = row[8].value

        if r + t + c + p == 100:
            e = leader(n, r, t, c, p, ps, ng)
            level6.append(e)
        elif me == 1 and meCheck is False:
            meCheck = True
        else:

            level6.append({"name": n, "error": "wrong values"})
            throwErr("total", f"{file}:{n}")



# returns a list of file names from a path
def get_files(directory):
    files = [f for f in listdir(directory) if isfile(join(directory, f))]
    return files

def throwErr(err = "", loc = ""):
    if err == "total":
        return print(f"ERROR: {loc} has not been totalled correctly")
    if err == "format":
        return print(f"ERROR: {loc} is not formatted correctly, only allows xslx, xsl, and ods")
    if err == "type":
        return print("ERROR: theres text in one of the cells or something. wtf?")
    return print("UNHANDLED: im pretty sure this wont ever get triggered")






if __name__ == '__main__':
    main()

